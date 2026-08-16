"""Structure-preserving Excel chessboard importer for MAYAK.

The importer deliberately keeps the Excel geometry separate from MAYAK's
business data.  Parsed workbooks are serialisable JSON dictionaries and can be
previewed before they are committed to ``data.json``.
"""

from copy import deepcopy
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter


FORMAT_VERSION = 2
DEFAULT_COLUMN_WIDTH = 96
DEFAULT_ROW_HEIGHT = 20
MIN_ROW_HEIGHT = 12
MAX_COLUMN_WIDTH = 220
MAX_ROW_HEIGHT = 160

APARTMENT_PATTERN = re.compile(
    r"(?:кв(?:артира)?\.?\s*№?\s*)(\d+[А-ЯA-Z]?)",
    re.IGNORECASE,
)
NON_RESIDENTIAL_PATTERN = re.compile(
    r"(?<!\d)(\d+\s*[-–—]\s*[НH])(?![А-ЯA-Z0-9])",
    re.IGNORECASE,
)
GENERIC_ROOM_PATTERN = re.compile(
    r"(?:пом(?:ещение)?\.?\s*№?\s*)([\wА-Яа-яЁё\-–—]+)",
    re.IGNORECASE,
)
ENTRANCE_PATTERN = re.compile(r"\b(парадн\w*|подъезд\w*|секци\w*)\b", re.IGNORECASE)
FLOOR_PATTERN = re.compile(
    r"(?:\b\d+\s*[-–—]?\s*этаж\w*\b|\bэтаж\w*\s*№?\s*\d+\b)",
    re.IGNORECASE,
)
BASEMENT_PATTERN = re.compile(r"\b(подвал\w*|цокол\w*)\b", re.IGNORECASE)
TECHNICAL_PATTERN = re.compile(
    r"\b(тех\.?\s*пом|техническ\w*|щитовая|колясочная|венткамера|моп|лестниц\w*)\b",
    re.IGNORECASE,
)
ADDRESS_PATTERN = re.compile(
    r"(?:\bул\.?\s|\bулиц\w*\s|\bпроспект\w*\s|\bд\.?\s*\d|\bдом\s*\d|\bжк\s+)",
    re.IGNORECASE,
)
WORK_TYPE_PATTERN = re.compile(
    r"\b(вид\s+работ|наименование\s+работ|капитальн\w*\s+ремонт|систем\w*|инженерн\w*\s+систем\w*)\b",
    re.IGNORECASE,
)
SERVICE_HEADER_PATTERN = re.compile(
    r"^(отч[её]т\s+о\s+выполненн\w*\s+работ\w*|адрес|объект|подрядчик|заказчик|этаж|помещение|квартира|во|хвс|гвс|тс)\s*:?.*$",
    re.IGNORECASE,
)


def cell_text(cell):
    """Return stable display text without converting empty cells to ``None``."""
    if cell.value is None:
        return ""

    if isinstance(cell.value, float) and cell.value.is_integer():
        return str(int(cell.value))

    return str(cell.value).strip()


def _normalise_room_number(value):
    return re.sub(r"\s*[-–—]\s*", "-", value.strip()).upper().replace("H", "Н")


def classify_cell(text):
    """Classify a cell using deterministic, conservative rules."""
    original = text.strip()
    lowered = original.lower()

    if not original:
        return {
            "cell_type": "empty_layout_cell",
            "room_number": None,
            "room_kind": None,
            "status": "white",
            "marking": "",
            "confidence": 1.0,
            "type": "empty",
        }

    apartment = APARTMENT_PATTERN.search(original)
    if apartment:
        number = apartment.group(1)
        return {
            "cell_type": "apartment",
            "room_number": number,
            "room_kind": "apartment",
            "status": "white",
            "marking": _extract_marking(original, apartment.group(0)),
            "confidence": 1.0,
            "type": "flat",
        }

    non_residential = NON_RESIDENTIAL_PATTERN.search(original)
    if non_residential:
        number = _normalise_room_number(non_residential.group(1))
        return {
            "cell_type": "non_residential",
            "room_number": number,
            "room_kind": "non_residential",
            "status": "white",
            "marking": _extract_marking(original, non_residential.group(0)),
            "confidence": 1.0,
            "type": "non_residential",
        }

    generic_room = GENERIC_ROOM_PATTERN.search(original)
    if generic_room:
        number = _normalise_room_number(generic_room.group(1))
        room_kind = "non_residential" if "нежил" in lowered else "room"
        return {
            "cell_type": "non_residential" if room_kind == "non_residential" else "technical_room",
            "room_number": number,
            "room_kind": room_kind,
            "status": "white",
            "marking": _extract_marking(original, generic_room.group(0)),
            "confidence": 0.85,
            "type": "non_residential" if room_kind == "non_residential" else "room",
        }

    if ENTRANCE_PATTERN.search(original):
        return _classification("entrance", 1.0)

    if FLOOR_PATTERN.search(original):
        return _classification("floor", 1.0)

    if BASEMENT_PATTERN.search(original):
        return _classification("basement", 1.0)

    if TECHNICAL_PATTERN.search(original):
        return _classification("technical_room", 0.9)

    if ADDRESS_PATTERN.search(original):
        return _classification("object_address", 0.85)

    if WORK_TYPE_PATTERN.search(original):
        return _classification("work_type", 0.9)

    if SERVICE_HEADER_PATTERN.match(original):
        return _classification("service_header", 0.8)

    return _classification("unknown", 0.25)


def _classification(cell_type, confidence):
    return {
        "cell_type": cell_type,
        "room_number": None,
        "room_kind": None,
        "status": "white",
        "marking": "",
        "confidence": confidence,
        "type": "other",
    }


def _extract_marking(text, matched_room):
    remainder = text.replace(matched_room, "", 1).replace("+", " ")
    return " ".join(remainder.split())


def _colour_to_hex(colour):
    if colour is None:
        return None

    if colour.type == "rgb" and colour.rgb:
        rgb = colour.rgb[-6:]
        if rgb != "000000" or colour.rgb.upper().endswith("000000"):
            return f"#{rgb.upper()}"

    if colour.type == "indexed" and colour.indexed is not None:
        index = int(colour.indexed)
        if 0 <= index < len(COLOR_INDEX):
            return f"#{COLOR_INDEX[index][-6:].upper()}"

    return None


def _border_side(side):
    if side is None:
        return {"style": None, "color": None}
    return {
        "style": side.style,
        "color": _colour_to_hex(side.color),
    }


def _cell_style(cell):
    fill = None
    if cell.fill and cell.fill.fill_type:
        fill = _colour_to_hex(cell.fill.fgColor) or _colour_to_hex(cell.fill.start_color)

    return {
        "fill": fill,
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
        "wrap_text": bool(cell.alignment.wrap_text),
        "rotation": int(cell.alignment.text_rotation or 0),
        "borders": {
            "top": _border_side(cell.border.top),
            "right": _border_side(cell.border.right),
            "bottom": _border_side(cell.border.bottom),
            "left": _border_side(cell.border.left),
        },
    }


def _has_visual_style(cell):
    if isinstance(cell, MergedCell):
        return any(
            getattr(side, "style", None)
            for side in (cell.border.top, cell.border.right, cell.border.bottom, cell.border.left)
        )

    if cell.fill and cell.fill.fill_type:
        return True

    if any(
        getattr(side, "style", None)
        for side in (cell.border.top, cell.border.right, cell.border.bottom, cell.border.left)
    ):
        return True

    return bool(cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical))


def _merge_maps(sheet):
    anchors = {}
    covered = set()

    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchors[(min_row, min_col)] = {
            "row_span": max_row - min_row + 1,
            "col_span": max_col - min_col + 1,
            "range": str(merged_range),
        }

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    covered.add((row, col))

    return anchors, covered


def _semantic_weight(text):
    cell_type = classify_cell(text)["cell_type"]
    if cell_type in {"apartment", "non_residential"}:
        return 20
    if cell_type == "entrance":
        return 10
    if cell_type in {"floor", "basement"}:
        return 6
    return 1


def _dominant_axis_cluster(cells, axis, maximum_gap):
    """Ignore isolated remote values without trimming genuine table gaps."""
    if not cells:
        return cells

    positions = sorted({getattr(cell, axis) for cell in cells})
    clusters = [[positions[0]]]
    for position in positions[1:]:
        if position - clusters[-1][-1] > maximum_gap:
            clusters.append([position])
        else:
            clusters[-1].append(position)

    if len(clusters) == 1:
        return cells

    scored = []
    for cluster in clusters:
        allowed = set(cluster)
        cluster_cells = [cell for cell in cells if getattr(cell, axis) in allowed]
        score = sum(_semantic_weight(cell_text(cell)) for cell in cluster_cells)
        scored.append((score, len(cluster_cells), cluster[-1] - cluster[0], cluster_cells))

    return max(scored, key=lambda item: item[:3])[3]


def _determine_working_bounds(sheet, anchors):
    materialised = list(sheet._cells.values())
    value_cells = [cell for cell in materialised if cell_text(cell)]
    value_cells = _dominant_axis_cluster(value_cells, "row", 12)
    value_cells = _dominant_axis_cluster(value_cells, "column", 8)
    value_coordinates = {
        (cell.row, cell.column)
        for cell in value_cells
    }
    structural_coordinates = {
        (cell.row, cell.column)
        for cell in materialised
        if cell_text(cell) or _has_visual_style(cell)
    }

    for (row, col), merge in anchors.items():
        anchor_cell = sheet.cell(row, col)
        anchor_text = cell_text(anchor_cell)
        selected_value = (row, col) in value_coordinates
        if anchor_text and not selected_value:
            continue
        if selected_value or _has_visual_style(anchor_cell):
            for merged_row in range(row, row + merge["row_span"]):
                for merged_col in range(col, col + merge["col_span"]):
                    structural_coordinates.add((merged_row, merged_col))
                    if selected_value:
                        value_coordinates.add((merged_row, merged_col))

    seed = value_coordinates or structural_coordinates
    if not seed:
        return 1, 1, 1, 1

    min_row = min(row for row, _ in seed)
    max_row = max(row for row, _ in seed)
    min_col = min(col for _, col in seed)
    max_col = max(col for _, col in seed)

    # A merged basement row is the semantic lower edge of a chessboard.  Some
    # workbooks retain white formatting in rows below that edge (for example,
    # from an old print area).  Rendering those style-only rows creates a fake
    # continuation of the board in MAYAK, although Excel shows the table ending
    # at the basement border.
    basement_bottoms = [
        row + merge["row_span"] - 1
        for (row, col), merge in anchors.items()
        if classify_cell(cell_text(sheet.cell(row, col)))["cell_type"] == "basement"
    ]
    semantic_bottom = max(basement_bottoms, default=None)
    if semantic_bottom is not None and semantic_bottom >= max_row:
        max_row = semantic_bottom

    # Include adjacent border/fill-only rows and columns only when they cover a
    # material part of the current table. Some source files keep sparse style
    # fragments below the real report; accepting any two styled cells would
    # turn those fragments into visible empty rows in MAYAK.
    def materially_continues_table(coordinate_count, axis_span):
        minimum = max(2, (axis_span * 2 + 4) // 5)
        return coordinate_count >= minimum

    changed = True
    while changed:
        changed = False

        top = [
            col for row, col in structural_coordinates
            if row == min_row - 1 and min_col <= col <= max_col
        ]
        if materially_continues_table(len(top), max_col - min_col + 1):
            min_row -= 1
            changed = True

        if semantic_bottom is None:
            bottom = [
                col for row, col in structural_coordinates
                if row == max_row + 1 and min_col <= col <= max_col
            ]
            if materially_continues_table(len(bottom), max_col - min_col + 1):
                max_row += 1
                changed = True

        left = [
            row for row, col in structural_coordinates
            if col == min_col - 1 and min_row <= row <= max_row
        ]
        if materially_continues_table(len(left), max_row - min_row + 1):
            min_col -= 1
            changed = True

        right = [
            row for row, col in structural_coordinates
            if col == max_col + 1 and min_row <= row <= max_row
        ]
        if materially_continues_table(len(right), max_row - min_row + 1):
            max_col += 1
            changed = True

    return min_row, max_row, min_col, max_col


def _column_width(sheet, col):
    dimension = sheet.column_dimensions[get_column_letter(col)]
    width = dimension.width
    if width is None:
        return DEFAULT_COLUMN_WIDTH
    return max(28, min(int(round(width * 7 + 5)), MAX_COLUMN_WIDTH))


def _row_height(sheet, row):
    height = sheet.row_dimensions[row].height
    if height is None:
        height = sheet.sheet_format.defaultRowHeight
    if height is None:
        return DEFAULT_ROW_HEIGHT
    return max(MIN_ROW_HEIGHT, min(int(round(height * 96 / 72)), MAX_ROW_HEIGHT))


def _worksheet_score(sheet):
    score = 0
    recognised_rooms = 0
    for cell in sheet._cells.values():
        text = cell_text(cell)
        if not text:
            continue
        classification = classify_cell(text)
        cell_type = classification["cell_type"]
        score += 1
        if cell_type in {"apartment", "non_residential"}:
            recognised_rooms += 1
            score += 20
        elif cell_type == "entrance":
            score += 10
        elif cell_type in {"floor", "basement"}:
            score += 6

    score += min(len(sheet.merged_cells.ranges), 100) * 0.5
    if sheet.sheet_state != "visible":
        score -= 1000
    return score, recognised_rooms


def _select_worksheet(workbook):
    scored = [(_worksheet_score(sheet), index, sheet) for index, sheet in enumerate(workbook.worksheets)]
    scored.sort(key=lambda item: (item[0][0], item[0][1], -item[1]), reverse=True)
    return scored[0][2], scored


def _context_for_cell(cell, entrance_cells, floor_cells):
    entrance = None
    entrance_candidates = [
        candidate for candidate in entrance_cells
        if candidate["row"] <= cell["row"]
        and candidate["col"] <= cell["col"] <= candidate["col"] + candidate["col_span"] - 1
    ]
    if entrance_candidates:
        entrance = max(entrance_candidates, key=lambda candidate: candidate["row"])["text"]

    floor = None
    floor_candidates = [
        candidate for candidate in floor_cells
        if candidate["row"] <= cell["row"] <= candidate["row"] + candidate["row_span"] - 1
    ]
    if floor_candidates:
        same_side = [candidate for candidate in floor_candidates if candidate["col"] <= cell["col"]]
        floor = (same_side or floor_candidates)[-1]["text"]

    return entrance, floor


def _column_interval(cell):
    return cell["col"], cell["col"] + cell["col_span"] - 1


def _interval_is_covered(start, end, intervals):
    """Return whether a set of column intervals continuously covers a slot."""
    cursor = start
    for interval_start, interval_end in sorted(intervals):
        if interval_end < cursor:
            continue
        if interval_start > cursor:
            return False
        cursor = max(cursor, interval_end + 1)
        if cursor > end:
            return True
    return False


def _is_contextual_room_label(cell, recognised_rooms):
    """Recognise a labelled room slot from workbook geometry.

    Some real chessboards identify non-residential rooms with a name rather
    than a number (for example, a shop, office or service tenant).  Such labels
    are valid room records when they occupy the same merged, bordered band as
    neighbouring apartments and belong to an entrance.  Geometry is used here
    deliberately; the rule is independent of any address or label text.
    """
    if (
        cell["cell_type"] != "unknown"
        or not cell["text"]
        or (not cell.get("entrance") and not cell.get("floor"))
    ):
        return False

    borders = cell["style"]["borders"]
    border_count = sum(1 for side in borders.values() if side.get("style"))
    has_horizontal_border = any(
        borders[side].get("style") for side in ("top", "bottom")
    )
    has_vertical_border = any(
        borders[side].get("style") for side in ("left", "right")
    )
    is_merged_slot = cell["row_span"] > 1 or cell["col_span"] > 1
    is_open_interior_slot = (
        border_count >= 2
        and has_horizontal_border
        and has_vertical_border
        and is_merged_slot
        and bool(cell.get("entrance"))
        and bool(cell.get("floor"))
    )
    if border_count < 3 and not is_open_interior_slot:
        return False

    def shares_room_band(room):
        if (
            room["row"] != cell["row"]
            or room["row_span"] != cell["row_span"]
            or room.get("floor") != cell.get("floor")
        ):
            return False

        if cell.get("entrance"):
            return room.get("entrance") == cell.get("entrance")

        # Some source sheets omit the final entrance header although the room
        # columns continue in the same bordered floor band.  In that case only
        # accept a slot that directly touches a recognised room horizontally;
        # a remote annotation elsewhere on the row must remain unknown.
        cell_start, cell_end = _column_interval(cell)
        room_start, room_end = _column_interval(room)
        return room_end + 1 == cell_start or cell_end + 1 == room_start

    same_room_band = any(shares_room_band(room) for room in recognised_rooms)
    if same_room_band:
        if border_count >= 3:
            return True

        # Interior cells in a horizontal room strip are sometimes drawn with
        # only the top and left separators.  The closing right/bottom borders
        # exist only on the final slot of the strip.  Accept such an open slot
        # only when it directly touches an already recognised room in exactly
        # the same entrance/floor band.  Fixed-point processing below lets a
        # chain of these slots grow from either recognised end without using
        # label text, workbook name, address or hard-coded coordinates.
        cell_start, cell_end = _column_interval(cell)
        return any(
            shares_room_band(room)
            and (
                _column_interval(room)[1] + 1 == cell_start
                or cell_end + 1 == _column_interval(room)[0]
            )
            for room in recognised_rooms
        )

    # Some construction matrices place a commercial slot in its own shallow
    # row immediately above or below the regular apartment band.  It therefore
    # has no floor label and a different row span, although its merged columns
    # are exactly covered by the neighbouring room cells.  Continuous column
    # coverage plus a shared entrance distinguishes such a room slot from a
    # remote note or an isolated styled label without relying on its wording.
    start_col, end_col = _column_interval(cell)
    cell_top = cell["row"]
    cell_bottom = cell["row"] + cell["row_span"] - 1
    adjacent_rooms = [
        room for room in recognised_rooms
        if room.get("entrance") == cell.get("entrance")
        and (
            room["row"] == cell_bottom + 1
            or room["row"] + room["row_span"] - 1 == cell_top - 1
        )
    ]
    adjacent_intervals = [
        _column_interval(room)
        for room in adjacent_rooms
        if _column_interval(room)[1] >= start_col
        and _column_interval(room)[0] <= end_col
    ]
    return _interval_is_covered(start_col, end_col, adjacent_intervals)


def _contextual_room_key(label, cell, used_room_keys):
    """Keep legacy label keys where possible and disambiguate duplicates."""
    if label not in used_room_keys:
        return label

    entrance = _normalise_room_number(cell.get("entrance") or "БЕЗ ПАРАДНОЙ")
    floor = _normalise_room_number(cell.get("floor") or "БЕЗ ЭТАЖА")
    source_slot = cell.get("merged_range") or cell.get("source_coordinate")
    structural_key = f"{label}::{entrance}::{floor}::{source_slot}"
    if structural_key not in used_room_keys:
        return structural_key

    # The source coordinate is unique within a worksheet.  This final branch
    # only protects malformed workbooks that repeat the same merged anchor.
    return f"{structural_key}::{cell['row']}:{cell['col']}"


def _parse_worksheet(sheet, workbook_sheet_names):
    anchors, covered = _merge_maps(sheet)
    min_row, max_row, min_col, max_col = _determine_working_bounds(sheet, anchors)

    columns = []
    for col in range(min_col, max_col + 1):
        width = _column_width(sheet, col)
        columns.append({
            "col": col,
            "letter": get_column_letter(col),
            "width": width,
            "hidden": bool(sheet.column_dimensions[get_column_letter(col)].hidden),
        })

    rows = []
    for row in range(min_row, max_row + 1):
        rows.append({
            "row": row,
            "height": _row_height(sheet, row),
            "hidden": bool(sheet.row_dimensions[row].hidden),
        })

    cells = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) in covered:
                continue

            cell = sheet.cell(row, col)
            text = cell_text(cell)
            classification = classify_cell(text)
            merge = anchors.get((row, col), {"row_span": 1, "col_span": 1, "range": None})
            row_span = merge["row_span"]
            col_span = merge["col_span"]
            width = sum(_column_width(sheet, item) for item in range(col, col + col_span))
            height = sum(_row_height(sheet, item) for item in range(row, row + row_span))
            hidden = any(
                sheet.row_dimensions[item].hidden
                for item in range(row, row + row_span)
            ) or any(
                sheet.column_dimensions[get_column_letter(item)].hidden
                for item in range(col, col + col_span)
            )

            parsed = {
                "row": row,
                "col": col,
                "grid_row": row - min_row + 1,
                "grid_col": col - min_col + 1,
                "text": text,
                "source_text": text,
                "row_span": row_span,
                "col_span": col_span,
                "merged_range": merge["range"],
                "source_coordinate": cell.coordinate,
                "width": width,
                "height": height,
                "hidden": bool(hidden),
                "entrance": None,
                "floor": None,
                "style": _cell_style(cell),
            }
            parsed.update(classification)
            parsed["number"] = parsed.get("room_number")
            parsed["room_key"] = parsed.get("room_number")
            cells.append(parsed)

    entrance_cells = [cell for cell in cells if cell["cell_type"] == "entrance"]
    floor_cells = [
        cell for cell in cells
        if cell["cell_type"] in {"floor", "basement"}
    ]

    for cell in cells:
        entrance, floor = _context_for_cell(cell, entrance_cells, floor_cells)
        if cell["cell_type"] != "entrance":
            cell["entrance"] = entrance
        if cell["cell_type"] not in {"floor", "basement"}:
            cell["floor"] = floor

    recognised_rooms = [
        cell for cell in cells
        if cell["cell_type"] in {"apartment", "non_residential"}
    ]
    used_room_keys = {
        str(cell.get("room_key") or cell.get("room_number") or "").strip()
        for cell in recognised_rooms
        if cell.get("room_key") or cell.get("room_number")
    }
    pending_contextual_cells = [
        cell for cell in cells
        if cell["cell_type"] == "unknown" and cell["text"]
    ]
    while pending_contextual_cells:
        remaining_cells = []
        recognised_in_pass = False

        for cell in pending_contextual_cells:
            if not _is_contextual_room_label(cell, recognised_rooms):
                remaining_cells.append(cell)
                continue

            room_number = _normalise_room_number(cell["text"])
            room_key = _contextual_room_key(room_number, cell, used_room_keys)
            cell.update({
                "cell_type": "non_residential",
                "room_number": room_number,
                "room_kind": "non_residential",
                "status": "white",
                "marking": "",
                "confidence": 0.7,
                "type": "non_residential",
                "number": room_number,
                "room_key": room_key,
            })
            used_room_keys.add(room_key)
            recognised_rooms.append(cell)
            recognised_in_pass = True

        if not recognised_in_pass:
            break
        pending_contextual_cells = remaining_cells

    rooms = [
        deepcopy(cell) for cell in cells
        if cell["cell_type"] in {"apartment", "non_residential"}
    ]
    unknown_cells = [
        cell for cell in cells
        if cell["cell_type"] == "unknown" and cell["text"]
    ]
    hidden_rows = [item["row"] for item in rows if item["hidden"]]
    hidden_columns = [item["col"] for item in columns if item["hidden"]]

    warnings = []
    if len(workbook_sheet_names) > 1:
        warnings.append(
            f"В книге найдено листов: {len(workbook_sheet_names)}. "
            f"Для импорта выбран лист «{sheet.title}»."
        )
    if unknown_cells:
        warnings.append(f"Не распознано заполненных ячеек: {len(unknown_cells)}.")
    if hidden_rows or hidden_columns:
        warnings.append(
            f"Найдены скрытые строки ({len(hidden_rows)}) "
            f"и столбцы ({len(hidden_columns)})."
        )
    if not rooms:
        warnings.append("Не найдено ни одного уверенно распознанного помещения.")

    return {
        "format_version": FORMAT_VERSION,
        "sheet_name": sheet.title,
        "sheet_names": list(workbook_sheet_names),
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "columns": columns,
        "rows": rows,
        "cells": cells,
        "rooms": rooms,
        "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_columns,
        "statistics": {
            "cells": len(cells),
            "rooms": len(rooms),
            "apartments": sum(1 for cell in rooms if cell["cell_type"] == "apartment"),
            "non_residential": sum(
                1 for cell in rooms if cell["cell_type"] == "non_residential"
            ),
            "entrances": len(entrance_cells),
            "floors": sum(1 for cell in floor_cells if cell["cell_type"] == "floor"),
            "basements": sum(1 for cell in floor_cells if cell["cell_type"] == "basement"),
            "unknown": len(unknown_cells),
            "merged": len(sheet.merged_cells.ranges),
        },
        "warnings": warnings,
    }


def parse_excel_chessboard(file_path):
    """Parse the most likely chessboard worksheet from an XLSX/XLSM file."""
    path = Path(file_path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet, _ = _select_worksheet(workbook)
    return _parse_worksheet(sheet, workbook.sheetnames)


def _legacy_cell(room, min_row, min_col):
    room = deepcopy(room)
    room_type = room.get("type", "other")
    cell_type = "apartment" if room_type == "flat" else "non_residential" if room_type in {
        "room", "office", "commercial"
    } else "unknown"
    room_number = str(room.get("room_number") or room.get("number") or "") or None
    status = room.get("status", "white")
    if status == "done":
        status = "green"
    elif status == "notdone":
        status = "white"

    room.update({
        "source_text": room.get("source_text", room.get("text", "")),
        "row_span": int(room.get("row_span", 1) or 1),
        "col_span": int(room.get("col_span", 1) or 1),
        "grid_row": int(room.get("row", min_row)) - min_row + 1,
        "grid_col": int(room.get("col", min_col)) - min_col + 1,
        "cell_type": room.get("cell_type", cell_type),
        "room_number": room_number,
        "room_kind": room.get("room_kind", "apartment" if cell_type == "apartment" else "non_residential"),
        "room_key": room.get("room_key", room_number),
        "status": status,
        "entrance": room.get("entrance"),
        "floor": room.get("floor"),
        "source_coordinate": room.get("source_coordinate"),
        "width": room.get("width"),
        "height": room.get("height"),
        "hidden": bool(room.get("hidden", False)),
        "confidence": float(room.get("confidence", 0.7)),
        "style": room.get("style", {}),
        "marking": room.get("marking", ""),
        "number": room_number,
    })
    return room


def normalize_chessboard(chessboard):
    """Return a version-2 view model for both new and legacy chessboards."""
    source = deepcopy(chessboard or {})
    min_row = int(source.get("min_row", 1) or 1)
    max_row = int(source.get("max_row", min_row) or min_row)
    min_col = int(source.get("min_col", 1) or 1)
    max_col = int(source.get("max_col", min_col) or min_col)

    if source.get("cells"):
        cells = []
        for item in source["cells"]:
            cell = deepcopy(item)
            cell.setdefault("row_span", 1)
            cell.setdefault("col_span", 1)
            cell.setdefault("grid_row", int(cell.get("row", min_row)) - min_row + 1)
            cell.setdefault("grid_col", int(cell.get("col", min_col)) - min_col + 1)
            cell.setdefault("cell_type", "unknown" if cell.get("text") else "empty_layout_cell")
            cell.setdefault("room_number", cell.get("number"))
            cell.setdefault("room_key", cell.get("room_number"))
            cell.setdefault("room_kind", None)
            cell.setdefault("status", "white")
            cell.setdefault("confidence", 0.5)
            cell["style"] = cell.get("style") or {}
            cell.setdefault("hidden", False)
            cell.setdefault("entrance", None)
            cell.setdefault("floor", None)
            cell.setdefault("marking", "")
            cells.append(cell)
    else:
        room_map = {
            (int(room.get("row", min_row)), int(room.get("col", min_col))): _legacy_cell(
                room, min_row, min_col
            )
            for room in source.get("rooms", [])
        }
        cells = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = room_map.get((row, col))
                if cell is None:
                    cell = {
                        "row": row,
                        "col": col,
                        "grid_row": row - min_row + 1,
                        "grid_col": col - min_col + 1,
                        "text": "",
                        "source_text": "",
                        "row_span": 1,
                        "col_span": 1,
                        "cell_type": "empty_layout_cell",
                        "room_number": None,
                        "room_key": None,
                        "room_kind": None,
                        "status": "white",
                        "confidence": 1.0,
                        "style": {},
                        "hidden": False,
                        "entrance": None,
                        "floor": None,
                        "marking": "",
                    }
                cells.append(cell)

    rooms = [
        deepcopy(cell) for cell in cells
        if cell.get("cell_type") in {"apartment", "non_residential"}
    ]
    columns = source.get("columns") or [
        {"col": col, "letter": get_column_letter(col), "width": DEFAULT_COLUMN_WIDTH, "hidden": False}
        for col in range(min_col, max_col + 1)
    ]
    rows = source.get("rows") or [
        {"row": row, "height": DEFAULT_ROW_HEIGHT, "hidden": False}
        for row in range(min_row, max_row + 1)
    ]
    unknown = sum(
        1 for cell in cells
        if cell.get("cell_type") == "unknown" and cell.get("text")
    )

    source.update({
        "format_version": int(source.get("format_version", 1)),
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "columns": columns,
        "rows": rows,
        "cells": cells,
        "rooms": rooms,
        "statistics": source.get("statistics") or {
            "cells": len(cells),
            "rooms": len(rooms),
            "apartments": sum(1 for cell in rooms if cell.get("cell_type") == "apartment"),
            "non_residential": sum(
                1 for cell in rooms if cell.get("cell_type") == "non_residential"
            ),
            "entrances": sum(1 for cell in cells if cell.get("cell_type") == "entrance"),
            "floors": sum(1 for cell in cells if cell.get("cell_type") == "floor"),
            "basements": sum(1 for cell in cells if cell.get("cell_type") == "basement"),
            "unknown": unknown,
            "merged": sum(1 for cell in cells if cell.get("row_span", 1) > 1 or cell.get("col_span", 1) > 1),
        },
        "warnings": list(source.get("warnings", [])),
    })
    return source


def apply_import_to_data(
    data,
    contractor,
    address,
    system,
    chessboard,
    existing_object=None,
    assigned_contractors=None,
):
    """Apply a confirmed preview to an isolated MAYAK data dictionary."""
    board = normalize_chessboard(chessboard)
    rooms = board.get("rooms", [])
    apartment_rooms = [room for room in rooms if room.get("cell_type") == "apartment"]
    floor_cells = [
        cell for cell in board.get("cells", [])
        if cell.get("cell_type") == "floor"
    ]
    floors_found = {
        room.get("floor") for room in rooms
        if room.get("floor")
    }

    existing_flats = int((existing_object or {}).get("flats", 1))
    existing_floors = int((existing_object or {}).get("floors", 1))
    flats = max(len(apartment_rooms) or existing_flats, 1)
    floors = max(len(floor_cells) or len(floors_found) or existing_floors, 1)
    systems = list((existing_object or {}).get("systems", []))

    if system not in systems:
        systems.append(system)

    contractor_names = list(assigned_contractors or [])
    if contractor not in contractor_names:
        contractor_names.append(contractor)

    imported_objects = data.setdefault("imported_objects", {})
    imported_objects[address] = {
        "systems": systems,
        "flats": flats,
        "floors": floors,
        "contractors": contractor_names,
    }

    key = f"{contractor}|{address}|{system}"
    data.setdefault("excel_chessboards", {})[key] = board
    data["last_excel_key"] = key

    system_data = data.setdefault(address, {}).setdefault(system, {})
    for room in rooms:
        room_key = str(room.get("room_key") or room.get("room_number") or "").strip()
        if not room_key or room_key in system_data:
            continue
        system_data[room_key] = {
            "status": "white",
            "contractor": contractor,
            "date": "",
            "photos": [],
        }

    return {
        "key": key,
        "flats": flats,
        "floors": floors,
        "systems": systems,
        "rooms": len(rooms),
    }
